from datetime import datetime
from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import *
import xlrd

now = datetime.now().isoformat()
Tk().withdraw()
p = filedialog.askopenfilename(initialdir = "/",title = "Выбор файла",filetypes = (("Файлы Excel *.xls*","*.xls*"),("Все файлы","*.*")))
if p == '' or p is None:
    print('Не выбран файл')
    exit()
f = open('ASFBKZKA_' + now[:4] + now[5:7] + now[8:10],'w',encoding='KZ-1048')
if p.split('.')[1] == 'xls':
    wb = xlrd.open_workbook(p,on_demand = True)
    ws = wb.sheet_by_index(0)
    for j in range(1,ws.nrows):
        for i in range(0,ws.ncols):
            if ws.cell(j,i).value == None or str(ws.cell(j,i).value) == '':
                continue
            else:
                s = str(ws.cell(j,i).value)
            if ws.cell(j,i).ctype == 3:
                s = xlrd.xldate.xldate_as_datetime(ws.cell(j,i).value,wb.datemode).isoformat()
                s = s[:4] + s[5:7] + s[8:10]
            if i > 0:
                f.write('@@')
            f.write(s)
        f.write('\n')
else:
    wb = load_workbook(p)
    ws = wb.active
    i = 0
    j = 0
    s = ''
    for row in ws.iter_rows():
        j += 1
        if j == 1:
            continue
        for cell in row:
            i += 1
            if i >= 1 and i <= 9:
                if cell.value == None:
                    continue
                else:
                    s = str(cell.value)
                if type(cell.value) is datetime:
                    s = s[:4] + s[5:7] + s[8:10]
                if i > 1:
                    f.write('@@')
                f.write(s)
        i = 0
        f.write('\n')
